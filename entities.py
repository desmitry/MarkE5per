from dataclasses import dataclass
from datetime import date
import pickle
@dataclass
class Mark:
    value: int
    date: date = None


@dataclass
class Subject:
    __slots__ = ['name', 'marks', 'average', 'goal']
    name: str
    marks: list
    average: float
    goal: int
    THRESHOLD = 0.71
    subjects = []
    marks_buffer = []


    def calculate_average(self):
        if self.marks:
            self.average = round(
                sum([mark.value for mark in self.marks]) / len(self.marks),
                2
            )


    def save():
        '''Save subjects in bytes.'''
        import pickle
        with open('subjects', 'wb') as f:
            pickle.dump(Subject.subjects, f)


    def load():
        '''Load subjects from bytes.'''
        import pickle
        with open('subjects', 'rb') as f:
            Subject.subjects = pickle.load(f)


    def load_excel(file):
        '''Load subjects from xlsx file.'''
        from openpyxl import load_workbook

        def create_mark(cell, value):
            month = None
            for i in range(cell.column, 0, -1):
                if sheet.cell(12, i).value:
                    month = MONTHS[sheet.cell(12, i).value]
                    break
            marks.append(
                Mark(
                    value=value,
                    date=date(
                        date.today().year,
                        # на винде B11. надо переделать так, чтобы не ссылаться так на вручную определенные номера строк
                        month,
                        # на винде 12
                        sheet.cell(13, cell.column).value
                    )
                )
            )

        MONTHS = {value:key for key, value in
            enumerate(
                ('Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль',
                    'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
                ),
                start=1
            )
        }
        sheet = load_workbook(filename=file).active
        Subject.subjects.clear()
        for row in sheet.iter_rows(14, 33, 1):
            marks = []
            for mark in row[1:]:
                if sheet.cell(12, mark.column).value == 'Средняя оценка':
                    break
                if not mark.value:
                    continue
                elif isinstance(mark.value, int):
                    create_mark(mark, mark.value)
                elif len(mark.value) > 1:
                    values = mark.value.split()
                    for value in values:
                        if value.isdigit():
                            create_mark(mark, int(value))
            subject = Subject(row[0].value, marks, None, 5)
            subject.calculate_average()
            Subject.subjects.append(subject)
            Subject.save()


    def return_remaining(self):
        """Return the amount of "5" (and "4") marks left for the average 
           of marks to satisfy the rounding threshold."""
        marks_sum = sum([mark.value for mark in self.marks])
        amount = len(self.marks)
        average = self.average
        fives_to_go = 0
        fours_to_go = 0
        while average < self.goal - 1 + Subject.THRESHOLD:
            fives_to_go += 1
            marks_sum += 5
            amount += 1
            average = (marks_sum) / (amount)
        average = self.average
        if self.goal != 5:
            while average < self.goal - 1 + Subject.THRESHOLD:
                fours_to_go += 1
                average = (marks_sum + 4) / (amount + 1)
        return fives_to_go, fours_to_go


if __name__ == '__main__': 
    answer = input('Загрузить данные из xlsx файла? (y/n): ')
    if answer == 'n':    
        Subject.load()
    else:
        Subject.load_excel('file2.xlsx')
        Subject.save()
    while True:
        eval(input('>>>'))
