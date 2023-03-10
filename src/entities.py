import os
from sys import platform
from dataclasses import dataclass
import datetime
import pickle

if platform == 'win32':
    SLASH = '\\'
elif platform == 'linux' or platform == 'darwin':
    SLASH = '/'
SCRIPT_PATH = os.path.dirname(os.path.realpath(__file__))

@dataclass(frozen=True)
class Mark:
    value: int
    date: datetime.date = datetime.date(1, 1, 1)
    is_backlog: bool = False

@dataclass
class Subject:
    __slots__ = ['name', 'marks', 'average', 'goal', 'to_add', 'to_remove']
    name: str
    marks: list
    average: float
    goal: int
    to_add: list
    to_remove: list
    threshold = 0.71
    subjects = []


    def calculate_average(self):
        if self.marks:
            self.average = round(
                sum([mark.value for mark in self.marks]) / len(self.marks),
                2
            )

    @classmethod
    def save(cls):
        '''Save subjects in bytes.'''
        with open(f'{SCRIPT_PATH}{SLASH}subjects', 'wb') as f:
            subjects = Subject.subjects.copy()
            subjects.append(Subject.threshold)
            pickle.dump(subjects, f)


    @classmethod
    def load(cls):
        '''Load subjects from bytes.'''
        with open(f'{SCRIPT_PATH}{SLASH}subjects', 'rb') as f:
            Subject.subjects = pickle.load(f)
            Subject.threshold = Subject.subjects.pop()


    @classmethod
    def load_excel(cls, file):
        '''Load subjects from xlsx file.'''
        if not file:
            return
        if platform == 'linux' or platform == 'darwin':
            MONTH_ROW = 12
        else:
            MONTH_ROW = 11
        DATE_ROW = MONTH_ROW + 1

        def sheet_from_xls(file):
            '''Return an openpyxl sheet with the contents of the specified .xls file.'''
            import xlrd
            from openpyxl.workbook import Workbook
            sheet1 = xlrd.open_workbook(file).sheet_by_index(0)
            nrows = sheet1.nrows
            ncols = sheet1.ncols
            sheet2 = Workbook.active
            for row in range(1, nrows):
                for column in range(1, ncols):
                    sheet2.cell(row=row, column=col).value = sheet2.cell_value(row, col)
            return sheet2

        def create_mark(cell, value):
            month = 0
            for i in range(cell.column, 0, -1):
                if sheet.cell(MONTH_ROW, i).value:
                    month = MONTHS[sheet.cell(MONTH_ROW, i).value]
                    break
            day = sheet.cell(DATE_ROW, cell.column).value
            if isinstance(value, int):
                subject.marks.append(
                    Mark(
                        value,
                        datetime.date(datetime.date.today().year, month, day)
                    )
                )
            else:
                subject.marks.append(
                    Mark(
                        2,
                        datetime.date(datetime.date.today().year, month, day),
                        is_backlog=True
                    )
                )
        MONTHS = {value:key for key, value in
            enumerate(
                ('????????????', '??????????????', '????????', '????????????', '??????', '????????', '????????',
                    '????????????', '????????????????', '??????????????', '????????????', '??????????????'
                ),
                start=1
            )
        }
        if file[-1] == 's':
            sheet = sheet_from_xls(file)
        else:
            from openpyxl import load_workbook
            sheet = load_workbook(filename=file).active
        Subject.subjects.clear()
        STOP_ROW = DATE_ROW
        STOP_COLUMN = 2
        while True:
            STOP_ROW += 1
            if not sheet.cell(STOP_ROW, 1).value:
                STOP_ROW -= 1
                break
        while True:
            STOP_COLUMN += 1
            if sheet.cell(MONTH_ROW, STOP_COLUMN).value == '?????????????? ????????????':
                STOP_COLUMN -= 1
                break
        for row in sheet.iter_rows(DATE_ROW + 1, STOP_ROW, 1, STOP_COLUMN):
            subject = Subject(row[0].value, [], 0, 5, [], [])
            for cell in row[1:]:
                if not cell.value:
                    continue
                elif isinstance(cell.value, int) or cell.value == '.':
                    create_mark(cell, cell.value)
                elif len(cell.value) > 1:
                    values = cell.value.split()
                    for value in values:
                        if value.isdigit():
                            create_mark(cell, int(value))
            subject.calculate_average()
            Subject.subjects.append(subject)


    def return_remaining(self):
        """Return the amount of "5" (and "4") marks left for the average
           of marks to satisfy the rounding threshold."""
        def calculate_to_go(value):
            marks_sum = sum([mark.value for mark in self.marks])
            amount = len(self.marks)
            average = self.average
            to_go = 0
            while average < self.goal - 1 + Subject.threshold:
                to_go += 1
                marks_sum += value
                amount += 1
                average = (marks_sum) / (amount)
            return to_go
        if self.goal == 5:
            return calculate_to_go(5), 0
        return calculate_to_go(5), calculate_to_go(4)


if __name__ == '__main__':
    answer = input('?????????????????? ???????????? ???? .xlsx ??????????? (Y/n) ')
    if answer == 'n':
        Subject.load()
    else:
        Subject.load_excel(str(input('?????????????? ???????? ?? ??????????:\n')))
        Subject.save()
    while True:
        eval(input('>>>'))
